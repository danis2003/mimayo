from openpyxl import load_workbook
from datetime import datetime
from scripts.config import HOJA_PRODUCTOS


def guardar_precios(productos, ruta_excel):

    libro = load_workbook(ruta_excel)
    hoja = libro[HOJA_PRODUCTOS]

    precios_modificados = 0
    precios_sin_cambios = 0

    for producto in productos:

        fila = producto["fila_excel"]

        precio_actual = hoja.cell(row=fila, column=5).value
        precio_nuevo = producto["precio"]

        if precio_actual != precio_nuevo:

            hoja.cell(row=fila, column=5).value = precio_nuevo

            celda_fecha = hoja.cell(row=fila, column=8)
            celda_fecha.value = datetime.now()
            celda_fecha.number_format = "dd/mm/yyyy"

            precios_modificados += 1

        else:

            precios_sin_cambios += 1

    try:

        libro.save(ruta_excel)

    except PermissionError:

        print("\n===================================")
        print(" ERROR")
        print("===================================\n")
        print("No se pudo guardar Excel_Maestro.xlsx.")
        print("Cierre el archivo en Excel e intente nuevamente.")

        return None, None

    return precios_modificados, precios_sin_cambios

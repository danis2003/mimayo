from openpyxl import load_workbook
from pathlib import Path
from scripts.config import (
    HOJA_PRODUCTOS,
    BASE_DIR,
    RUTA_ICONO,
    RUTA_CONFIG
)
import tkinter as tk
from tkinter import filedialog, messagebox
# Pillow nos permite abrir, redimensionar y mostrar imágenes.
from PIL import Image, ImageTk, ImageFile
# Permite abrir PNG ligeramente truncados o incompletos
ImageFile.LOAD_TRUNCATED_IMAGES = True
import shutil
import json
import re
from scripts.generar_json import main as generar_json

# ==========================================
# RUTAS
# ==========================================

EXCEL = BASE_DIR / "data" / "Excel_Maestro.xlsx"

CARPETA_PENDIENTES = BASE_DIR / "img" / "pendientes"

CARPETA_PRODUCTOS = BASE_DIR / "img" / "productos"
# ==========================================
# CONFIGURACIÓN DE VENTANA
# ==========================================

RUTA_VENTANA_ASISTENTE = (
    RUTA_CONFIG / "asistente_imagenes_ventana.json"
)

# ==========================================
# VARIABLES GLOBALES
# ==========================================

# Imagen utilizada en el flujo normal
archivo_seleccionado = None

# Imágenes seleccionadas para agregar variantes
archivos_variantes = []

# Producto seleccionado para agregar variantes

producto_variantes = None
# Producto actualmente administrado mediante el CRUD
producto_administrar = None

modo_administracion = False

# Orden original de las imágenes al cargar el producto
archivos_variantes_originales = []

# Imágenes que el producto ya tenía antes de agregar nuevas variantes
archivos_variantes_existentes = []

# Indica si estamos trabajando en modo administración


# Imagen que se encuentra seleccionada para reemplazo
indice_imagen_reemplazo = None

# Nueva imagen seleccionada para reemplazar la imagen existente
imagen_nueva_reemplazo = None

# Imágenes marcadas para eliminación.
# La eliminación física se realiza únicamente
# al presionar "Guardar cambios".
imagenes_marcadas_eliminar = []

# ==========================================
# CARGAR EXCEL
# ==========================================

wb = load_workbook(EXCEL)

ws = wb[HOJA_PRODUCTOS]

# ==========================================
# BUSCAR EL PRIMER PRODUCTO SIN IMAGEN
# Devuelve el primer producto pendiente junto
# con su posición y la cantidad total.
# ==========================================

def obtener_producto_pendiente():

    total_productos = ws.max_row - 1
    numero_producto = 0

    # Recorremos todas las filas del Excel
    for fila in range(2, ws.max_row + 1):

        numero_producto += 1

        imagen = ws[fila][5].value

        # Si la columna Imagen está vacía,
        # encontramos el siguiente producto.
        if imagen is None or str(imagen).strip() == "":

            return {
                "fila": fila,
                "numero": numero_producto,
                "total": total_productos,
                "codigo": ws[fila][0].value,
                "nombre": ws[fila][1].value,
                "marca": ws[fila][2].value,
                "categoria": ws[fila][3].value,
            }

    return None
# ==========================================
# MOSTRAR EL PRODUCTO EN LA INTERFAZ
# Actualiza los Labels con la información
# del producto encontrado.
# ==========================================

def mostrar_producto():

    producto = obtener_producto_pendiente()

    if producto is None:

        lblProducto.config(text="¡No quedan productos pendientes!")

        lblCodigo.config(text="-")
        lblNombre.config(text="-")
        lblMarca.config(text="-")
        lblCategoria.config(text="-")

        return

    lblProducto.config(
        text=f"Producto {producto['numero']} de {producto['total']}"
    )

    lblCodigo.config(text=producto["codigo"])
    lblNombre.config(text=producto["nombre"])
    lblMarca.config(text=producto["marca"])
    lblCategoria.config(text=producto["categoria"])


# ==========================================
# CARGAR IMÁGENES EXISTENTES DEL PRODUCTO
# Solo lectura.
# No modifica archivos.
# ==========================================
def cargar_imagenes_producto(producto):
    VARIANTES_JSON = BASE_DIR / "data" / "variantes.json"

    codigo = str(producto["codigo"]).strip()

    imagenes = []

    # ------------------------------------------
    # Imagen principal desde el Excel
    # ------------------------------------------
    imagen_principal = producto.get("imagen")

    if imagen_principal:
        nombre_principal = str(imagen_principal).strip()

        if nombre_principal:
            ruta_principal = CARPETA_PRODUCTOS / nombre_principal

            if ruta_principal.exists():
                imagenes.append(ruta_principal)

    # ------------------------------------------
    # Variantes desde variantes.json
    # ------------------------------------------
    if VARIANTES_JSON.exists():
        try:
            with open(
                VARIANTES_JSON,
                "r",
                encoding="utf-8"
            ) as archivo:
                variantes = json.load(archivo)

            nombres_variantes = variantes.get(codigo, [])

            for nombre in nombres_variantes:
                ruta = CARPETA_PRODUCTOS / str(nombre).strip()

                if ruta.exists():
                    imagenes.append(ruta)

        except (json.JSONDecodeError, OSError):
            pass

    return imagenes


# ==========================================
# SELECCIONAR PRODUCTO PARA VARIANTES Y ADMINISTRAR IMAGENES
# Permite elegir un producto existente del Excel.
# ==========================================

def seleccionar_producto_variantes(modo="variantes"):

    global producto_variantes
    global archivos_variantes
    global archivo_seleccionado
    global archivos_variantes_existentes

    ventana = tk.Toplevel(root)
    ventana.title("Seleccionar producto")
    ventana.geometry("900x650")
    ventana.minsize(750, 550)
    ventana.configure(bg="white")

    # ------------------------------------------
    # TÍTULO
    # ------------------------------------------

    tk.Label(
        ventana,
        text="Seleccione el producto",
        font=("Segoe UI", 18, "bold"),
        bg="white"
    ).pack(
        pady=(20, 10)
    )

    # ------------------------------------------
    # BUSCADOR
    # ------------------------------------------

    tk.Label(
        ventana,
        text="Buscar por código o nombre:",
        font=("Segoe UI", 10, "bold"),
        bg="white"
    ).pack(
        anchor="w",
        padx=25
    )

    entrada_busqueda = tk.Entry(
        ventana,
        font=("Segoe UI", 12)
    )

    entrada_busqueda.pack(
        fill="x",
        padx=25,
        pady=(5, 10)
    )

    # ------------------------------------------
    # CONTENEDOR DE LISTA + SCROLLBAR
    # ------------------------------------------

    frameLista = tk.Frame(
        ventana,
        bg="white"
    )

    frameLista.pack(
        fill="both",
        expand=True,
        padx=25,
        pady=5
    )

    scrollbar = tk.Scrollbar(
        frameLista
    )

    scrollbar.pack(
        side="right",
        fill="y"
    )

    lista = tk.Listbox(
        frameLista,
        font=("Segoe UI", 11),
        selectmode=tk.SINGLE,
        yscrollcommand=scrollbar.set
    )

    lista.pack(
        side="left",
        fill="both",
        expand=True
    )

    scrollbar.config(
        command=lista.yview
    )

    # ------------------------------------------
    # CARGAR PRODUCTOS
    # ------------------------------------------

    productos = []

    for fila in range(2, ws.max_row + 1):

        codigo = ws[fila][0].value
        nombre = ws[fila][1].value
        marca = ws[fila][2].value
        categoria = ws[fila][3].value
        imagen = ws[fila][5].value

        if codigo is not None and nombre is not None:

            productos.append({
                "fila": fila,
                "codigo": codigo,
                "nombre": nombre,
                "marca": marca,
                "categoria": categoria,
                "imagen": imagen
            })

    # ------------------------------------------
    # ORDENAR POR CÓDIGO ASCENDENTE
    # ------------------------------------------

    def clave_codigo(producto):

        codigo = str(producto["codigo"]).strip()

        try:
            return (0, int(codigo))
        except ValueError:
            return (1, codigo.lower())

    productos.sort(
        key=clave_codigo
    )

    # ------------------------------------------
    # ACTUALIZAR LISTA SEGÚN BÚSQUEDA
    # ------------------------------------------

    productos_filtrados = []

    def actualizar_lista(event=None):

        nonlocal productos_filtrados

        texto = entrada_busqueda.get().strip().lower()

        productos_filtrados = [
            producto
            for producto in productos
            if texto in str(producto["codigo"]).lower()
            or texto in str(producto["nombre"]).lower()
        ]

        lista.delete(
            0,
            "end"
        )

        for producto in productos_filtrados:

            lista.insert(
                "end",
                f"{producto['codigo']} - {producto['nombre']}"
            )

    entrada_busqueda.bind(
        "<KeyRelease>",
        actualizar_lista
    )

    actualizar_lista()

    # ------------------------------------------
    # CONFIRMAR PRODUCTO
    # ------------------------------------------

    def confirmar_producto():

        global producto_variantes
        global archivos_variantes
        global archivo_seleccionado

        seleccion = lista.curselection()

        if not seleccion:

            messagebox.showwarning(
                "Producto",
                "Seleccione un producto."
            )

            return

        producto = productos_filtrados[
            seleccion[0]
        ]

        ventana.destroy()

        # ======================================
        # MODO AGREGAR VARIANTES
        # ======================================

        if modo == "variantes":

            global archivos_variantes_existentes

            producto_variantes = producto
            archivos_variantes = []
            archivos_variantes_existentes = cargar_imagenes_producto(producto)
            archivo_seleccionado = None

            actualizar_estado(
                f"Producto seleccionado: "
                f"{producto['codigo']} - "
                f"{producto['nombre']}"
            )

            lblCodigo.config(
                text=producto["codigo"]
            )

            lblNombre.config(
                text=producto["nombre"]
            )

            lblMarca.config(
                text=producto["marca"]
            )

            lblCategoria.config(
                text=producto["categoria"]
            )

            limpiar_previsualizacion()

            if archivos_variantes_existentes:
                mostrar_previsualizacion_variantes()

            return

        # if modo == "variantes":

        #     producto_variantes = producto
        #     archivos_variantes = []
        #     archivo_seleccionado = None

        #     actualizar_estado(
        #         f"Producto seleccionado: "
        #         f"{producto['codigo']} - "
        #         f"{producto['nombre']}"
        #     )

        #     lblCodigo.config(
        #         text=producto["codigo"]
        #     )

        #     lblNombre.config(
        #         text=producto["nombre"]
        #     )

        #     lblMarca.config(
        #         text=producto["marca"]
        #     )

        #     lblCategoria.config(
        #         text=producto["categoria"]
        #     )

        #     return

        # ======================================
        # MODO ADMINISTRAR IMÁGENES
        # ======================================

        if modo == "administrar":

            global producto_administrar
            global archivos_variantes_originales
            global modo_administracion

            producto_variantes = None
            archivo_seleccionado = None

            producto_administrar = producto
            modo_administracion = True

            # ------------------------------------------
            # Cargar imágenes existentes
            # ------------------------------------------

            archivos_variantes = cargar_imagenes_producto(
                producto
            )

            # ------------------------------------------
            # Guardar una copia del orden original
            # ------------------------------------------

            archivos_variantes_originales = list(
                archivos_variantes
            )

            # ------------------------------------------
            # Mostrar información del producto
            # ------------------------------------------

            lblProducto.config(
                text="Administrando imágenes"
            )

            lblCodigo.config(
                text=producto["codigo"]
            )

            lblNombre.config(
                text=producto["nombre"]
            )

            lblMarca.config(
                text=producto["marca"]
            )

            lblCategoria.config(
                text=producto["categoria"]
            )

            # ------------------------------------------
            # En Administración NO usamos Confirmar
            # ------------------------------------------

            btnConfirmar.config(
                state="disabled"
            )

            btnGuardarCambios.config(
                state="disabled"
            )

            limpiar_previsualizacion()

            # ------------------------------------------
            # Mostrar imágenes existentes
            # ------------------------------------------

            if archivos_variantes:

                mostrar_previsualizacion_variantes()

                actualizar_estado(
                    f"{len(archivos_variantes)} imagen(es) "
                    "cargada(s) para administrar."
                )

            else:

                actualizar_estado(
                    "El producto no tiene imágenes cargadas."
                )

    # ------------------------------------------
    # BOTÓN
    # ------------------------------------------

    tk.Button(
        ventana,
        text="Seleccionar producto",
        font=("Segoe UI", 11, "bold"),
        bg="#2563EB",
        fg="white",
        padx=20,
        pady=10,
        cursor="hand2",
        command=confirmar_producto
    ).pack(
        pady=(10, 20)
    )


# ==========================================
# ADMINISTRAR IMÁGENES
# Inicia el flujo de lectura de imágenes
# existentes de un producto.
#
# Este primer paso es SOLO READ:
# no modifica archivos ni JSON.
# ==========================================

def administrar_imagenes():

    global archivos_variantes
    global archivos_variantes_originales
    global archivo_seleccionado
    global producto_variantes
    global producto_administrar
    global modo_administracion
    global indice_imagen_reemplazo
    global imagen_nueva_reemplazo
    global imagenes_marcadas_eliminar
    global archivos_variantes_existentes

    # ------------------------------------------
    # Preparar modo administración
    # ------------------------------------------

    archivos_variantes = []
    archivos_variantes_originales = []
    archivos_variantes_existentes = []

    archivo_seleccionado = None
    producto_variantes = None
    producto_administrar = None

    modo_administracion = True

    limpiar_previsualizacion()

    btnConfirmar.config(
        state="disabled"
    )

    btnGuardarCambios.config(
        state="disabled"
    )

    actualizar_estado(
        "Seleccione el producto que desea administrar."
    )

    seleccionar_producto_variantes(
        modo="administrar"
    )


# ==========================================
# PREVISUALIZAR REEMPLAZO DE IMAGEN
#
# Muestra temporalmente la nueva imagen
# seleccionada para reemplazar una imagen
# existente.
#
# No modifica todavía ningún archivo físico.
# ==========================================

def mostrar_previsualizacion_reemplazo(
    indice,
    imagen_tk,
    imagen_nueva
):

    # ------------------------------------------
    # Limpiar la previsualización actual
    # ------------------------------------------

    for widget in frameVariantes.winfo_children():
        widget.destroy()

    if not archivos_variantes:
        return

    # ------------------------------------------
    # Contenedor de imágenes
    # ------------------------------------------

    frameImagenes = tk.Frame(
        frameVariantes,
        bg="#F5F5F5"
    )

    frameImagenes.pack(
        fill="x"
    )

    # ------------------------------------------
    # Crear nuevamente las tarjetas
    # ------------------------------------------

    for i, ruta in enumerate(archivos_variantes):

        frameImagen = tk.Frame(
            frameImagenes,
            bg="#DBEAFE" if i == indice else "#F3F4F6",
            bd=2,
            relief="solid"
        )

        frameImagen.grid(
            row=0,
            column=i,
            padx=5,
            pady=5,
            sticky="n"
        )

        # --------------------------------------
        # Imagen
        # --------------------------------------

        if i == indice:

            imagen_mostrar = imagen_tk

        else:

            try:

                imagen = Image.open(ruta)

                imagen.thumbnail(
                    (150, 150)
                )

                imagen_mostrar = ImageTk.PhotoImage(
                    imagen
                )

                frameImagen.imagen = imagen_mostrar

            except Exception:

                imagen_mostrar = None

        if imagen_mostrar:

            labelImagen = tk.Label(
                frameImagen,
                image=imagen_mostrar,
                bg="#DBEAFE" if i == indice else "#F3F4F6"
            )

            labelImagen.image = imagen_mostrar

            labelImagen.pack(
                padx=8,
                pady=8
            )

        # --------------------------------------
        # Nombre
        # --------------------------------------

        if i == indice:

            nombre = Path(
                imagen_nueva
            ).name

        else:

            nombre = Path(
                ruta
            ).name

        tk.Label(
            frameImagen,
            text=nombre,
            font=("Segoe UI", 9, "bold"),
            bg="#DBEAFE" if i == indice else "#F3F4F6"
        ).pack(
            pady=(0, 4)
        )

        # --------------------------------------
        # Indicador
        # --------------------------------------

        if i == indice:

            tk.Label(
                frameImagen,
                text="NUEVA IMAGEN",
                font=("Segoe UI", 9, "bold"),
                bg="#DBEAFE",
                fg="#1D4ED8"
            ).pack(
                fill="x",
                padx=8,
                pady=(0, 6)
            )

        elif i == 0:

            tk.Label(
                frameImagen,
                text="PRINCIPAL",
                font=("Segoe UI", 9, "bold"),
                bg="#DCFCE7",
                fg="#166534"
            ).pack(
                fill="x",
                padx=8,
                pady=(0, 6)
            )

        # --------------------------------------
        # Permitir seleccionar otra imagen
        # --------------------------------------

        def seleccionar_para_reemplazo(
            evento=None,
            indice_actual=i
        ):

            global indice_imagen_reemplazo

            if not modo_administracion:
                return

            indice_imagen_reemplazo = indice_actual

            mostrar_previsualizacion_variantes()

            actualizar_estado(
                f"Imagen seleccionada para reemplazo: "
                f"{Path(archivos_variantes[indice_actual]).name}"
            )

        frameImagen.bind(
            "<Button-1>",
            seleccionar_para_reemplazo
        )

        if 'labelImagen' in locals():

            labelImagen.bind(
                "<Button-1>",
                seleccionar_para_reemplazo
            )

            
# ==========================================
# SELECCIONAR IMAGEN
# Abre el explorador y muestra una vista previa.
# No guarda nada todavía.
# ==========================================

def seleccionar_imagen():

    global archivo_seleccionado
    global archivos_variantes

    global archivos_variantes_existentes

    # ==========================================
    # MODO ADMINISTRACIÓN
    # ==========================================

    if modo_administracion:

        # --------------------------------------
        # Primero debemos tener una imagen
        # existente seleccionada.
        # --------------------------------------

        if indice_imagen_reemplazo is None:

            messagebox.showwarning(
                "Reemplazar imagen",
                "Primero seleccione la imagen existente "
                "que desea reemplazar."
            )

            return

        # --------------------------------------
        # Seleccionar nueva imagen
        # --------------------------------------

        archivos = filedialog.askopenfilenames(
            title="Seleccionar nueva imagen",
            filetypes=[
                (
                    "Imágenes",
                    "*.png *.jpg *.jpeg *.webp *.bmp *.gif"
                )
            ]
        )

        if not archivos:
            return

        # --------------------------------------
        # Para reemplazar una imagen solamente
        # aceptamos UNA imagen.
        # --------------------------------------

        if len(archivos) > 1:

            messagebox.showwarning(
                "Reemplazar imagen",
                "Para reemplazar una imagen existente "
                "seleccione solamente una imagen."
            )

            return

        imagen_nueva_reemplazo = Path(
            archivos[0]
        )

        # --------------------------------------
        # Procesar temporalmente la nueva imagen
        # para obtener una previsualización.
        #
        # IMPORTANTE:
        # todavía NO reemplazamos el archivo
        # existente.
        # --------------------------------------

        try:

            imagen = Image.open(
                imagen_nueva_reemplazo
            )

            imagen.thumbnail(
                (300, 300)
            )

            imagen_tk = ImageTk.PhotoImage(
                imagen
            )

            # ----------------------------------
            # Mostrar la nueva imagen
            # en la posición seleccionada.
            # ----------------------------------

            ruta_actual = archivos_variantes[
                indice_imagen_reemplazo
            ]

            # Guardamos temporalmente la imagen
            # nueva para la previsualización.

            # archivos_variantes_preview = list(
            #     archivos_variantes
            # )

            # Esta lista es solamente visual.
            # No modifica todavía archivos_variantes.

            mostrar_previsualizacion_reemplazo(
                indice_imagen_reemplazo,
                imagen_tk,
                imagen_nueva_reemplazo
            )

            actualizar_estado(
                f"Nueva imagen preparada para reemplazar: "
                f"{Path(ruta_actual).name}"
            )

            btnGuardarCambios.config(
                state="normal"
            )

        except Exception as error:

            imagen_nueva_reemplazo = None

            messagebox.showerror(
                "Error",
                f"No se pudo cargar la nueva imagen.\n\n"
                f"{error}"
            )

        return

    # ==========================================
    # MODO NORMAL
    # ==========================================

    archivos = filedialog.askopenfilenames(
        title="Seleccionar imagen(es)",
        filetypes=[
            (
                "Imágenes",
                "*.png *.jpg *.jpeg *.webp *.bmp *.gif"
            )
        ]
    )

    if not archivos:
        return

    # ==========================================
    # AGREGAR VARIANTES A PRODUCTO EXISTENTE
    # ==========================================

    if producto_variantes is not None:

        archivos_variantes.extend(
            Path(archivo)
            for archivo in archivos
        )

        mostrar_previsualizacion_variantes()

        actualizar_estado(
            f"{len(archivos_variantes)} nueva(s) imagen(es) "
            "seleccionada(s)."
        )

        btnConfirmar.config(
            state="normal"
        )

        return

    # ------------------------------------------
    # Una sola imagen
    # ------------------------------------------

    if len(archivos) == 1:

        archivo_seleccionado = Path(
            archivos[0]
        )

        archivos_variantes = [
            archivo_seleccionado
        ]

        mostrar_vista_previa(
            archivo_seleccionado
        )

        actualizar_estado(
            "1 imagen seleccionada."
        )

        btnConfirmar.config(
            state="normal"
        )

        return

    # if len(archivos) == 1:

    #     archivo_seleccionado = Path(
    #         archivos[0]
    #     )

    #     # procesar_imagen(
    #     #     archivo_seleccionado
    #     # )

    #     return

    # ------------------------------------------
    # Varias imágenes
    # ------------------------------------------

    archivos_variantes = [
        Path(archivo)
        for archivo in archivos
    ]

    mostrar_previsualizacion_variantes()

    actualizar_estado(
        f"{len(archivos_variantes)} imágenes "
        "seleccionadas."
    )

    btnConfirmar.config(
        state="normal"
    )

# ==========================================
# MOSTRAR VISTA PREVIA
# Carga la imagen elegida y la muestra
# dentro del recuadro.
# ==========================================

def mostrar_vista_previa(ruta):

    # Abrimos la imagen
    imagen = Image.open(ruta)

    # Mantiene la proporción
    imagen.thumbnail((420,420))

    # Conversión para Tkinter
    foto = ImageTk.PhotoImage(imagen)

    # Mostrar la imagen
    lblImagen.config(
        image=foto,
        text=""
    )

    # Evita que Python elimine la imagen
    lblImagen.image = foto

# ==========================================
# MOVER IMAGEN EN LA PREVISUALIZACIÓN
# Cambia la posición de una imagen dentro
# de la lista de imágenes seleccionadas.
# ==========================================

def mover_imagen(indice, desplazamiento):

    global archivos_variantes

    nueva_posicion = indice + desplazamiento

    if nueva_posicion < 0 or nueva_posicion >= len(archivos_variantes):
        return

    # ------------------------------------------
    # Intercambiar posiciones solamente
    # en memoria.
    # ------------------------------------------

    archivos_variantes[indice], archivos_variantes[nueva_posicion] = (
        archivos_variantes[nueva_posicion],
        archivos_variantes[indice]
    )

    # ------------------------------------------
    # Si estamos administrando un producto,
    # determinar si existen cambios pendientes.
    # ------------------------------------------

    if modo_administracion:

        if archivos_variantes != archivos_variantes_originales:

            btnGuardarCambios.config(
                state="normal"
            )

        else:

            btnGuardarCambios.config(
                state="disabled"
            )

    # ------------------------------------------
    # Reconstruir previsualización
    # ------------------------------------------

    mostrar_previsualizacion_variantes()


# ==========================================
# MARCAR IMAGEN PARA ELIMINACIÓN
#
# La imagen NO se elimina físicamente todavía.
# Se quita de la previsualización y se registra
# para que "Guardar cambios" ejecute la operación.
# ==========================================

def marcar_imagen_para_eliminar(indice):

    global archivos_variantes
    global indice_imagen_reemplazo
    global imagen_nueva_reemplazo
    global imagenes_marcadas_eliminar

    if not modo_administracion:
        return

    if indice < 0 or indice >= len(archivos_variantes):
        return

    ruta = Path(
        archivos_variantes[indice]
    )

    # ------------------------------------------
    # Evitar duplicados
    # ------------------------------------------

    if ruta not in imagenes_marcadas_eliminar:
        imagenes_marcadas_eliminar.append(ruta)

    # ------------------------------------------
    # Si esta era la imagen seleccionada para
    # reemplazo, cancelar ese reemplazo.
    # ------------------------------------------

    if indice_imagen_reemplazo == indice:

        indice_imagen_reemplazo = None
        imagen_nueva_reemplazo = None

    elif (
        indice_imagen_reemplazo is not None
        and indice_imagen_reemplazo > indice
    ):

        # Al quitar una imagen, los índices
        # posteriores se desplazan una posición.
        indice_imagen_reemplazo -= 1

    # ------------------------------------------
    # Quitar solamente de la lista en memoria.
    # ------------------------------------------

    archivos_variantes.pop(indice)

    # ------------------------------------------
    # Determinar estado
    # ------------------------------------------

    if archivos_variantes:

        actualizar_estado(
            f"Imagen marcada para eliminar: "
            f"{ruta.name}"
        )

        mostrar_previsualizacion_variantes()

    else:

        limpiar_previsualizacion()

        actualizar_estado(
            f"Imagen marcada para eliminar: "
            f"{ruta.name}. "
            "El producto quedará sin imágenes."
        )

    # ------------------------------------------
    # Hay cambios pendientes.
    # ------------------------------------------

    btnGuardarCambios.config(
        state="normal"
    )

# ==========================================
# GUARDAR CAMBIOS DE IMÁGENES
#
# Ejecuta las modificaciones realizadas
# durante la administración del producto.
#
# Actualmente contempla:
# - Reemplazo de una imagen existente.
# - Eliminación de una o varias imágenes.
# - Reordenamiento de imágenes.
#
# El reemplazo NO utiliza procesar_imagen()
# porque esa función mueve la imagen original
# a /procesadas_png.
#
# El CRUD procesa la nueva imagen directamente
# en /productos y envía la imagen anterior
# a /papelera.
#
# Las eliminaciones mueven las imágenes
# eliminadas a /papelera y reorganizan
# las imágenes restantes.
#
# También actualiza:
# - variantes.json
# - Excel Maestro
# - productos.json
#
# La operación utiliza un sistema de temporales
# para permitir la restauración ante errores.
# ==========================================

def guardar_cambios_imagenes():

    global archivos_variantes
    global archivos_variantes_originales
    global producto_administrar
    global modo_administracion
    global imagen_nueva_reemplazo
    global indice_imagen_reemplazo
    global imagenes_marcadas_eliminar

    # ------------------------------------------
    # VALIDACIONES
    # ------------------------------------------

    if not modo_administracion:

        messagebox.showwarning(
            "Administración",
            "No hay un producto en modo administración."
        )

        return

    if producto_administrar is None:

        messagebox.showwarning(
            "Administración",
            "No hay ningún producto seleccionado."
        )

        return

    # ------------------------------------------
    # Detectar cambios
    # ------------------------------------------

    hay_reemplazo = (
        indice_imagen_reemplazo is not None
        and imagen_nueva_reemplazo is not None
    )

    hay_eliminaciones = bool(
        imagenes_marcadas_eliminar
    )

    hay_cambio_orden = (
        not hay_eliminaciones
        and archivos_variantes != archivos_variantes_originales
    )

    if (
        not hay_reemplazo
        and not hay_eliminaciones
        and not hay_cambio_orden
    ):

        messagebox.showinfo(
            "Sin cambios",
            "No hay cambios para guardar."
        )

        btnGuardarCambios.config(
            state="disabled"
        )

        return

    # ------------------------------------------
    # Confirmación
    # ------------------------------------------

    cambios = []

    if hay_reemplazo:
        cambios.append("reemplazo de imagen")

    if hay_cambio_orden:
        cambios.append("cambio de orden")

    if hay_eliminaciones:
        cambios.append("eliminación de imagen(es)")

    descripcion = "\n".join(
        f"• {cambio}"
        for cambio in cambios
    )

    respuesta = messagebox.askyesno(
        "Guardar cambios",
        "Se realizarán los siguientes cambios:\n\n"
        f"{descripcion}\n\n"
        "¿Desea continuar?"
    )

    if not respuesta:
        return

    # ------------------------------------------
    # Carpetas de trabajo
    # ------------------------------------------

    carpeta_papelera = (
        BASE_DIR / "img" / "papelera"
    )

    carpeta_temporal = (
        BASE_DIR / "img" / ".tmp_crud"
    )

    carpeta_papelera.mkdir(
        parents=True,
        exist_ok=True
    )

    carpeta_temporal.mkdir(
        parents=True,
        exist_ok=True
    )

    # Guardamos información para poder restaurar
    # si ocurre un error.
    movimientos_transaccion = []
    try:

        # ======================================
        # DELETE
        # ELIMINAR IMÁGENES
        # ======================================

        if hay_eliminaciones:

            # ----------------------------------
            # Por seguridad, no combinar todavía
            # eliminación y reemplazo en una misma
            # operación.
            # ----------------------------------

            if hay_reemplazo:

                raise RuntimeError(
                    "No se puede combinar un reemplazo "
                    "y una eliminación en la misma "
                    "operación. Guarde primero una "
                    "operación y luego la otra."
                )

            # ----------------------------------
            # Preparar listas
            # ----------------------------------

            originales = [
                Path(ruta)
                for ruta in archivos_variantes_originales
            ]

            actuales = [
                Path(ruta)
                for ruta in archivos_variantes
            ]

            eliminadas = {
                Path(ruta)
                for ruta in imagenes_marcadas_eliminar
            }

            # ----------------------------------
            # Verificar que las imágenes marcadas
            # realmente pertenecen al producto.
            # ----------------------------------

            if not eliminadas.issubset(
                set(originales)
            ):

                raise RuntimeError(
                    "Una o más imágenes marcadas "
                    "para eliminar no pertenecen "
                    "al producto administrado."
                )

            # ----------------------------------
            # Verificar que todas las imágenes
            # originales existan.
            # ----------------------------------

            for ruta in originales:

                if not ruta.exists():

                    raise FileNotFoundError(
                        "No se encontró la imagen:\n\n"
                        f"{ruta}"
                    )

            # ----------------------------------
            # Mover TODAS las imágenes originales
            # a temporales.
            #
            # Esto permite reorganizar nombres
            # sin colisiones.
            # ----------------------------------

            mapa_temporales = {}

            for indice, ruta in enumerate(
                originales
            ):

                temporal = (
                    carpeta_temporal
                    / f".tmp_delete_{indice}_"
                    f"{ruta.name}"
                )

                contador = 1

                while temporal.exists():

                    temporal = (
                        carpeta_temporal
                        / f".tmp_delete_{indice}_"
                        f"{contador}_"
                        f"{ruta.name}"
                    )

                    contador += 1

                shutil.move(
                    str(ruta),
                    str(temporal)
                )

                mapa_temporales[
                    str(ruta)
                ] = temporal

                movimientos_transaccion.append(
                    {
                        "original": ruta,
                        "actual": temporal
                    }
                )

            # ----------------------------------
            # Enviar las imágenes eliminadas
            # a papelera.
            # ----------------------------------

            for ruta in eliminadas:

                temporal = mapa_temporales[
                    str(ruta)
                ]

                ruta_papelera = (
                    carpeta_papelera
                    / ruta.name
                )

                contador = 1

                while ruta_papelera.exists():

                    ruta_papelera = (
                        carpeta_papelera
                        / f"{ruta.stem}_"
                        f"eliminada_{contador}"
                        f"{ruta.suffix}"
                    )

                    contador += 1

                shutil.move(
                    str(temporal),
                    str(ruta_papelera)
                )

                for movimiento in movimientos_transaccion:

                    if movimiento["actual"] == temporal:

                        movimiento["actual"] = ruta_papelera

                        break

            # ----------------------------------
            # Reubicar las imágenes restantes
            # según su nuevo orden.
            #
            # El primer archivo siempre vuelve
            # a ser codigo.webp.
            # ----------------------------------

            for indice, ruta_origen in enumerate(
                actuales
            ):

                temporal = mapa_temporales[
                    str(ruta_origen)
                ]

                ruta_destino = originales[
                    indice
                ]

                shutil.move(
                    str(temporal),
                    str(ruta_destino)
                )

                for movimiento in movimientos_transaccion:

                    if movimiento["actual"] == temporal:

                        movimiento["actual"] = ruta_destino

                        break

            # ----------------------------------
            # Verificar resultado.
            # ----------------------------------

            for indice in range(
                len(actuales)
            ):

                ruta_destino = originales[
                    indice
                ]

                if not ruta_destino.exists():

                    raise RuntimeError(
                        "No se pudo verificar "
                        f"el archivo {ruta_destino.name}"
                    )

            # ----------------------------------
            # Actualizar variantes.json
            # ----------------------------------

            VARIANTES_JSON = (
                BASE_DIR
                / "data"
                / "variantes.json"
            )

            if VARIANTES_JSON.exists():

                try:

                    with open(
                        VARIANTES_JSON,
                        "r",
                        encoding="utf-8"
                    ) as archivo:

                        variantes = json.load(
                            archivo
                        )

                except (
                    json.JSONDecodeError,
                    OSError
                ):

                    variantes = {}

            else:

                variantes = {}

            codigo = str(
                producto_administrar["codigo"]
            ).strip()

            nuevos_nombres_variantes = [
                ruta.name
                for ruta in originales[
                    1:len(actuales)
                ]
            ]

            if nuevos_nombres_variantes:

                variantes[codigo] = (
                    nuevos_nombres_variantes
                )

            else:

                variantes.pop(
                    codigo,
                    None
                )

            # ----------------------------------
            # Guardar variantes.json
            # ----------------------------------

            with open(
                VARIANTES_JSON,
                "w",
                encoding="utf-8"
            ) as archivo:

                json.dump(
                    variantes,
                    archivo,
                    indent=4,
                    ensure_ascii=False
                )

            # ----------------------------------
            # Actualizar Excel Maestro
            # ----------------------------------

            if actuales:

                nuevo_nombre_principal = (
                    f"{codigo}.webp"
                )

                ws[
                    producto_administrar["fila"]
                ][5].value = (
                    nuevo_nombre_principal
                )

            else:

                # El producto quedó sin imagen.
                ws[
                    producto_administrar["fila"]
                ][5].value = None

            wb.save(EXCEL)

        # ======================================
        # CASO 1
        # REEMPLAZAR UNA IMAGEN
        # ======================================

        if hay_reemplazo:

            ruta_original = Path(
                archivos_variantes_originales[
                    indice_imagen_reemplazo
                ]
            )

            nueva_imagen = Path(
                imagen_nueva_reemplazo
            )

            # ----------------------------------
            # Verificar original
            # ----------------------------------

            if not ruta_original.exists():

                raise FileNotFoundError(
                    "No se encontró la imagen original:\n\n"
                    f"{ruta_original}"
                )

            # ----------------------------------
            # Verificar nueva imagen
            # ----------------------------------

            if not nueva_imagen.exists():

                raise FileNotFoundError(
                    "No se encontró la nueva imagen:\n\n"
                    f"{nueva_imagen}"
                )

            # ----------------------------------
            # Crear nombre temporal único
            # ----------------------------------

            temporal = (
                carpeta_temporal /
                f".tmp_{ruta_original.name}"
            )

            contador = 1

            while temporal.exists():

                temporal = (
                    carpeta_temporal /
                    f".tmp_{contador}_"
                    f"{ruta_original.name}"
                )

                contador += 1

            # ----------------------------------
            # PASO 1
            #
            # Imagen original → temporal
            # ----------------------------------

            shutil.move(
                str(ruta_original),
                str(temporal)
            )

            movimientos_transaccion.append(
                {
                    "original": ruta_original,
                    "actual": temporal
                }
            )

            # ----------------------------------
            # PASO 2
            #
            # Procesar nueva imagen directamente
            # en el nombre original.
            # ----------------------------------

            imagen = Image.open(
                nueva_imagen
            )

            try:

                # PNG con paleta
                if imagen.mode == "P":

                    imagen = imagen.convert(
                        "RGBA"
                    )

                # Otros modos incompatibles
                elif imagen.mode not in (
                    "RGB",
                    "RGBA"
                ):

                    imagen = imagen.convert(
                        "RGBA"
                    )

                imagen.save(
                    ruta_original,
                    format="WEBP",
                    quality=90,
                    method=0
                )

            finally:

                imagen.close()

            # ----------------------------------
            # PASO 3
            #
            # Verificar nueva imagen
            # ----------------------------------

            if not ruta_original.exists():

                raise RuntimeError(
                    "La nueva imagen no pudo "
                    "ser creada correctamente."
                )

            # ----------------------------------
            # PASO 4
            #
            # Mover imagen vieja → papelera
            # ----------------------------------

            ruta_papelera = (
                carpeta_papelera /
                ruta_original.name
            )

            contador = 1

            while ruta_papelera.exists():

                ruta_papelera = (
                    carpeta_papelera /
                    f"{ruta_original.stem}_"
                    f"eliminada_{contador}"
                    f"{ruta_original.suffix}"
                )

                contador += 1

            shutil.move(
                str(temporal),
                str(ruta_papelera)
            )

            for movimiento in movimientos_transaccion:

                if movimiento["actual"] == temporal:

                    movimiento["actual"] = ruta_papelera

                    break

            # Ya no necesitamos restaurar este
            # temporal porque la operación terminó.
            movimientos_transaccion.clear()

        # ======================================
        # CASO 2
        # CAMBIO DE ORDEN
        # ======================================

        if hay_cambio_orden:

            # ----------------------------------
            # Verificar cantidad
            # ----------------------------------

            if len(archivos_variantes) != len(
                archivos_variantes_originales
            ):

                raise RuntimeError(
                    "La cantidad de imágenes cambió "
                    "inesperadamente."
                )

            # ----------------------------------
            # Verificar que sean las mismas
            # imágenes.
            # ----------------------------------

            if set(archivos_variantes) != set(
                archivos_variantes_originales
            ):

                raise RuntimeError(
                    "Las imágenes administradas "
                    "no coinciden con las originales."
                )

            # ----------------------------------
            # Verificar existencia
            # ----------------------------------

            for ruta in archivos_variantes_originales:

                ruta = Path(ruta)

                if not ruta.exists():

                    raise FileNotFoundError(
                        "No se encontró la imagen:\n\n"
                        f"{ruta}"
                    )

            # ----------------------------------
            # Crear temporales
            # ----------------------------------

            mapa_temporales = {}

            for indice, ruta in enumerate(
                archivos_variantes_originales
            ):

                ruta = Path(ruta)

                temporal = (
                    carpeta_temporal /
                    f".tmp_orden_{indice}_"
                    f"{ruta.name}"
                )

                contador = 1

                while temporal.exists():

                    temporal = (
                        carpeta_temporal /
                        f".tmp_orden_{indice}_"
                        f"{contador}_"
                        f"{ruta.name}"
                    )

                    contador += 1

                shutil.move(
                    str(ruta),
                    str(temporal)
                )

                movimientos_transaccion.append(
                    {
                        "original": ruta,
                        "actual": temporal
                    }
                )

                mapa_temporales[
                    str(ruta)
                ] = temporal

            # ----------------------------------
            # Colocar cada imagen en su nuevo
            # nombre.
            # ----------------------------------

            for indice, ruta_destino in enumerate(
                archivos_variantes_originales
            ):

                ruta_origen = Path(
                    archivos_variantes[indice]
                )

                temporal = mapa_temporales[
                    str(ruta_origen)
                ]

                shutil.move(
                    str(temporal),
                    str(ruta_destino)
                )

                for movimiento in movimientos_transaccion:

                    if movimiento["actual"] == temporal:

                        movimiento["actual"] = ruta_destino

                        break

            # ----------------------------------
            # Verificar resultado
            # ----------------------------------

            for ruta in archivos_variantes_originales:

                ruta = Path(ruta)

                if not ruta.exists():

                    raise RuntimeError(
                        "No se pudo verificar "
                        f"el archivo {ruta.name}"
                    )

        # ======================================
        # ÉXITO
        # ======================================

        if hay_eliminaciones:

            # Después de eliminar/reordenar, las rutas
            # físicas definitivas son las de "originales".

            archivos_variantes = list(
                originales[:len(actuales)]
            )

            archivos_variantes_originales = list(
                archivos_variantes
            )

            imagen_nueva_reemplazo = None
            indice_imagen_reemplazo = None
            imagenes_marcadas_eliminar = []

            btnGuardarCambios.config(
                state="disabled"
            )

            actualizar_estado(
                "Cambios guardados correctamente."
            )

            mostrar_previsualizacion_variantes()

            generar_json()

            # Si el producto quedó sin imágenes,
            # salir del modo administración.
            if not actuales:

                modo_administracion = False
                producto_administrar = None

                archivos_variantes = []
                archivos_variantes_originales = []

                imagen_nueva_reemplazo = None
                indice_imagen_reemplazo = None
                imagenes_marcadas_eliminar = []

                btnGuardarCambios.config(state="disabled")
                btnConfirmar.config(state="disabled")

                limpiar_previsualizacion()

                mostrar_producto()

                actualizar_estado(
                    "Seleccione una imagen para comenzar."
                )

            messagebox.showinfo(
                "Cambios guardados",
                "Las modificaciones se guardaron "
                "correctamente."
            )

    except Exception as error:

        # ======================================
        # RESTAURACIÓN
        # ======================================

        try:

            for movimiento in reversed(
                movimientos_transaccion
            ):

                ruta_original = movimiento[
                    "original"
                ]

                ruta_actual = movimiento[
                    "actual"
                ]

                # ----------------------------------
                # Si el archivo ya está en su ubicación
                # original, no hay nada que restaurar.
                # ----------------------------------

                if ruta_actual == ruta_original:

                    continue

                # ----------------------------------
                # Verificar que el archivo exista
                # actualmente.
                # ----------------------------------

                if not ruta_actual.exists():

                    raise FileNotFoundError(
                        "No se encontró el archivo "
                        "durante la restauración:\n\n"
                        f"{ruta_actual}"
                    )

                # ----------------------------------
                # Si la ubicación original está ocupada,
                # eliminar el archivo que quedó allí
                # durante la operación fallida.
                # ----------------------------------

                if ruta_original.exists():

                    try:

                        ruta_original.unlink()

                    except Exception as error_eliminar_original:

                        raise RuntimeError(
                            "No se pudo preparar la ubicación "
                            "original para la restauración:\n\n"
                            f"{ruta_original}"
                        ) from error_eliminar_original

                # ----------------------------------
                # Restaurar el archivo a su ubicación
                # original.
                #
                # Puede ser:
                #
                # papelera → productos
                # .tmp_crud → productos
                # productos → productos
                # ----------------------------------

                shutil.move(
                    str(ruta_actual),
                    str(ruta_original)
                )

        except Exception as error_restauracion:

            messagebox.showerror(
                "Error crítico",
                "No se pudo completar la operación "
                "y tampoco se pudo restaurar "
                "completamente el estado anterior.\n\n"
                f"Error original:\n{error}\n\n"
                f"Error de restauración:\n"
                f"{error_restauracion}"
            )

            return

        messagebox.showerror(
            "Error",
            "No se pudieron guardar los cambios.\n\n"
            "Se intentó restaurar el estado anterior.\n\n"
            f"Detalle:\n{error}"
        )

    finally:

        # --------------------------------------
        # Limpiar archivos temporales restantes
        # --------------------------------------

        try:

            if carpeta_temporal.exists():

                for archivo in carpeta_temporal.iterdir():

                    try:
                        archivo.unlink()
                    except Exception:
                        pass

                try:
                    carpeta_temporal.rmdir()
                except Exception:
                    pass

        except Exception:
            pass


# ==========================================
# PREVISUALIZAR VARIANTES
# Muestra miniaturas de todas las imágenes
# seleccionadas y permite cambiar su orden.
# ==========================================

def mostrar_previsualizacion_variantes():

    global indice_imagen_reemplazo
    global archivos_variantes_existentes

    # ------------------------------------------
    # Limpiar contenido anterior
    # ------------------------------------------

    for widget in frameVariantes.winfo_children():
        widget.destroy()

    if not archivos_variantes and not archivos_variantes_existentes:
        return

    # ------------------------------------------
    # Ocultar la vista previa individual.
    #
    # Cuando existen imágenes administradas,
    # las miniaturas ocupan el área principal.
    # ------------------------------------------

    lblImagen.pack_forget()

    # ------------------------------------------
    # Preparar contenedor de variantes
    # ------------------------------------------

    frameVariantes.pack_forget()

    frameVariantes.pack(
        fill="both",
        expand=True,
        padx=10,
        pady=10
    )

    # ------------------------------------------
    # CONTENEDOR CON SCROLL
    # ------------------------------------------

    frameScroll = tk.Frame(
        frameVariantes,
        bg="#F5F5F5"
    )

    frameScroll.pack(
        fill="both",
        expand=True
    )

    canvas = tk.Canvas(
        frameScroll,
        bg="#F5F5F5",
        highlightthickness=0
    )

    scrollbar = tk.Scrollbar(
        frameScroll,
        orient="vertical",
        command=canvas.yview
    )

    canvas.configure(
        yscrollcommand=scrollbar.set
    )

    scrollbar.pack(
        side="right",
        fill="y"
    )

    canvas.pack(
        side="left",
        fill="both",
        expand=True
    )

    frameImagenes = tk.Frame(
        canvas,
        bg="#F5F5F5"
    )

    canvas.create_window(
        (0, 0),
        window=frameImagenes,
        anchor="nw"
    )

    frameImagenes.bind(
        "<Configure>",
        lambda evento: canvas.configure(
            scrollregion=canvas.bbox("all")
        )
    )

    canvas.bind(
        "<Configure>",
        lambda evento: canvas.itemconfig(
            canvas.find_withtag("all")[0],
            width=evento.width
        )
    )

    # Scroll con rueda del mouse
    def scroll_rueda(evento):
        canvas.yview_scroll(
            int(-1 * (evento.delta / 120)),
            "units"
        )

    canvas.bind(
        "<Enter>",
        lambda evento: canvas.bind_all(
            "<MouseWheel>",
            scroll_rueda
        )
    )

    canvas.bind(
        "<Leave>",
        lambda evento: canvas.unbind_all(
            "<MouseWheel>"
        )
    )

    # ------------------------------------------
    # Configurar cuadrícula
    # ------------------------------------------

    for columna in range(3):
        frameImagenes.grid_columnconfigure(
            columna,
            weight=1
        )

    # ------------------------------------------
    # CREAR MINIATURAS
    # ------------------------------------------

        if producto_variantes is not None:
            imagenes_mostrar = (
                archivos_variantes_existentes
                + archivos_variantes
            )
        else:
            imagenes_mostrar = archivos_variantes

    for indice, ruta in enumerate(imagenes_mostrar):

        frameImagen = tk.Frame(
            frameImagenes,
            bg="#F3F4F6",
            bd=2,
            relief="solid",
            cursor="hand2"
        )

        frameImagen.grid(
            row=indice // 3,
            column=indice % 3,
            padx=5,
            pady=5,
            sticky="n"
        )

        # --------------------------------------
        # Imagen
        # --------------------------------------

        try:

            imagen = Image.open(ruta)

            imagen.thumbnail(
                (150, 150)
            )

            imagen_tk = ImageTk.PhotoImage(
                imagen
            )

            labelImagen = tk.Label(
                frameImagen,
                image=imagen_tk,
                bg="#F3F4F6",
                cursor="hand2"
            )

            labelImagen.image = imagen_tk

            labelImagen.pack(
                padx=8,
                pady=8
            )

        except Exception:

            tk.Label(
                frameImagen,
                text="No se pudo\ncargar",
                font=("Segoe UI", 10),
                bg="#F3F4F6",
                width=18,
                height=8
            ).pack(
                padx=8,
                pady=8
            )

        # --------------------------------------
        # Número / nombre
        # --------------------------------------

        nombre = Path(ruta).name

        tk.Label(
            frameImagen,
            text=nombre,
            font=("Segoe UI", 9, "bold"),
            bg="#F3F4F6"
        ).pack(
            pady=(0, 4)
        )

        # --------------------------------------
        # Identificar principal
        # --------------------------------------

        if indice == 0:

            tk.Label(
                frameImagen,
                text="PRINCIPAL",
                font=("Segoe UI", 9, "bold"),
                bg="#DCFCE7",
                fg="#166534"
            ).pack(
                fill="x",
                padx=8,
                pady=(0, 6)
            )

        # --------------------------------------
        # Seleccionar para reemplazo
        # --------------------------------------

        def seleccionar_para_reemplazo(
            evento=None,
            indice_actual=indice
        ):

            global indice_imagen_reemplazo

            if not modo_administracion:
                return

            indice_imagen_reemplazo = indice_actual

            # Actualizar apariencia de todas
            # las miniaturas.

            for widget in frameImagenes.winfo_children():

                widget.configure(
                    bg="#F3F4F6"
                )

                for hijo in widget.winfo_children():

                    try:
                        hijo.configure(
                            bg="#F3F4F6"
                        )
                    except Exception:
                        pass

            frameImagen.configure(
                bg="#DBEAFE"
            )

            for hijo in frameImagen.winfo_children():

                try:
                    hijo.configure(
                        bg="#DBEAFE"
                    )
                except Exception:
                    pass

            nombre_imagen_seleccionada = Path(
                archivos_variantes[indice_actual]
            ).name

            actualizar_estado(
                f"Imagen seleccionada para reemplazo: "
                f"{nombre_imagen_seleccionada}"
            )

        # --------------------------------------
        # Vincular clic a toda la tarjeta
        # --------------------------------------

        frameImagen.bind(
            "<Button-1>",
            seleccionar_para_reemplazo
        )

        labelImagen.bind(
            "<Button-1>",
            seleccionar_para_reemplazo
        )

        # --------------------------------------
        # Controles de movimiento
        # --------------------------------------

        frameBotones = tk.Frame(
            frameImagen,
            bg="#F3F4F6"
        )

        frameBotones.pack(
            pady=(0, 8)
        )

        btnArriba = tk.Button(
            frameBotones,
            text="▲",
            width=3,
            command=lambda i=indice: mover_imagen(
                i,
                -1
            )
        )

        btnArriba.pack(
            side="left",
            padx=2
        )

        btnAbajo = tk.Button(
            frameBotones,
            text="▼",
            width=3,
            command=lambda i=indice: mover_imagen(
                i,
                1
            )
        )

        btnAbajo.pack(
            side="left",
            padx=2
        )

        # --------------------------------------
        # Botón eliminar
        # --------------------------------------

        btnEliminar = tk.Button(
            frameImagen,
            text="🗑 Eliminar",
            font=("Segoe UI", 9, "bold"),
            fg="#B91C1C",
            cursor="hand2",
            command=lambda i=indice: marcar_imagen_para_eliminar(i)
        )

        btnEliminar.pack(
            pady=(0, 8)
        )

# ==========================================
# Limpiar imagen principal
# ==========================================
def limpiar_previsualizacion():

    # ------------------------------------------
    # Limpiar imagen principal
    # ------------------------------------------

    lblImagen.config(
        image="",
        text="Vista previa\n\n(Sin imagen)"
    )

    lblImagen.image = None

    # ------------------------------------------
    # Limpiar miniaturas
    # ------------------------------------------

    for widget in frameVariantes.winfo_children():
        widget.destroy()

    # ------------------------------------------
    # Restaurar distribución original
    # ------------------------------------------

    frameVariantes.pack_forget()

    lblImagen.pack(
        expand=True
    )

    frameVariantes.pack(
        fill="x",
        padx=10,
        pady=10
    )
# ==========================================
# FUNCION PARA CAMBIAR FORMATO DE IMAGENES A WEBP Y MOVERLAS A /PROCESADAS_PNG
# ==========================================
def procesar_imagen(archivo, nuevo_nombre):
    
    destino = CARPETA_PRODUCTOS / nuevo_nombre

    # --------------------------------------
    # Abrir la imagen
    # --------------------------------------

    imagen = Image.open(archivo)

    # --------------------------------------
    # Convertir al modo adecuado
    # --------------------------------------

    # PNG con paleta → RGBA para conservar transparencia
    if imagen.mode == "P":
        imagen = imagen.convert("RGBA")

    # Otros modos → RGB/RGBA compatible con WEBP
    elif imagen.mode not in ("RGB", "RGBA"):
        imagen = imagen.convert("RGBA")

    # --------------------------------------
    # Guardar como WEBP
    # --------------------------------------

    imagen.save(
        destino,
        format="WEBP",
        quality=90,
        method=0
    )

    imagen.close()

    # --------------------------------------
    # Verificar y mover original
    # --------------------------------------

    if destino.exists():

        carpeta_procesadas = (
            BASE_DIR / "img" / "procesadas_png"
        )

        carpeta_procesadas.mkdir(
            exist_ok=True
        )

        shutil.move(
            archivo,
            carpeta_procesadas / Path(archivo).name
        )

        return True

    return False
# ==========================================
# GUARDAR IMAGEN
# Convierte la imagen a WEBP, la renombra,
# la mueve a la carpeta productos y
# actualiza el Excel.
# ==========================================

def guardar_imagen():

    global archivo_seleccionado

    if archivo_seleccionado is None:
        return False

    producto = obtener_producto_pendiente()

    if producto is None:
        return False

    # --------------------------------------
    # Nombre final del archivo
    # --------------------------------------

    nuevo_nombre = f"{producto['codigo']}.webp"

    destino = CARPETA_PRODUCTOS / nuevo_nombre

    # --------------------------------------
    # Si ya existe, preguntar si se reemplaza
    # --------------------------------------

    if destino.exists():

        respuesta = messagebox.askyesno(
            "Imagen existente",
            f"Ya existe la imagen:\n\n{nuevo_nombre}\n\n¿Desea reemplazarla?"
        )

        if not respuesta:
            return False

    # --------------------------------------
    # Procesar imagen
    # --------------------------------------

    if not procesar_imagen(
        archivo_seleccionado,
        nuevo_nombre
    ):
        messagebox.showerror(
            "Error",
            "No se pudo guardar la imagen."
        )
        return False
    # --------------------------------------
    # Actualizar Excel
    # --------------------------------------

    ws[producto["fila"]][5].value = nuevo_nombre
    wb.save(EXCEL)
    actualizar_estado(f"Imagen guardada: {nuevo_nombre}")
    
    return True

# ==========================================
# función que guarda las variantes
# ==========================================
def guardar_variantes(producto, archivos):

    if not archivos:
        return True

    VARIANTES_JSON = BASE_DIR / "data" / "variantes.json"

    # ------------------------------------------
    # Cargar variantes existentes
    # ------------------------------------------

    if VARIANTES_JSON.exists():

        try:
            with open(
                VARIANTES_JSON,
                "r",
                encoding="utf-8"
            ) as archivo:

                variantes = json.load(archivo)

        except (json.JSONDecodeError, OSError):

            variantes = {}

    else:
        variantes = {}

    codigo = str(producto["codigo"])

    if codigo not in variantes:
        variantes[codigo] = []

    # ------------------------------------------
    # Determinar siguiente número
    # ------------------------------------------

    existentes = variantes[codigo]

    # ------------------------------------------
    # Determinar próximo número de variante
    # ------------------------------------------

    numeros = []

    patron = re.compile(
        rf"^{re.escape(codigo)}-(\d+)\.webp$"
    )

    for nombre in existentes:

        coincidencia = patron.match(nombre)

        if coincidencia:
            numeros.append(
                int(coincidencia.group(1))
            )

    siguiente = max(numeros, default=1) + 1

    # ------------------------------------------
    # Procesar imágenes
    # ------------------------------------------

    for archivo in archivos:

        nuevo_nombre = (
            f"{codigo}-{siguiente}.webp"
        )

        if procesar_imagen(
            archivo,
            nuevo_nombre
        ):

            variantes[codigo].append(
                nuevo_nombre
            )

            siguiente += 1


    # ------------------------------------------
    # Guardar variantes.json
    # ------------------------------------------

    with open(
        VARIANTES_JSON,
        "w",
        encoding="utf-8"
    ) as archivo:

        json.dump(
            variantes,
            archivo,
            indent=4,
            ensure_ascii=False
        )

    return True
# ==========================================
# CONFIRMAR IMAGEN
# Guarda la imagen y avanza al siguiente
# producto.
# ==========================================

def confirmar_imagen():

    global archivo_seleccionado
    global archivos_variantes
    global producto_variantes
    global archivos_variantes_existentes

    # ------------------------------------------
    # VALIDAR QUE HAYA IMÁGENES
    # ------------------------------------------

    if not archivos_variantes:
        messagebox.showwarning(
            "Sin imágenes",
            "Seleccione al menos una imagen."
        )
        return

    # ------------------------------------------
    # PRODUCTO EXISTENTE → AGREGAR VARIANTES
    # ------------------------------------------

    if producto_variantes is not None:

        if guardar_variantes(
            producto_variantes,
            archivos_variantes
        ):
            actualizar_estado(
                "Variantes guardadas correctamente."
            )

            producto_variantes = None
            archivo_seleccionado = None
            archivos_variantes = []
            archivos_variantes_existentes = []

            btnConfirmar.config(
                state="disabled"
            )

            limpiar_previsualizacion()
            mostrar_producto()

        return

    # ------------------------------------------
    # PRODUCTO SIN IMAGEN
    # ------------------------------------------

    producto = obtener_producto_pendiente()

    if producto is None:
        return

    # La primera imagen es la principal
    archivo_seleccionado = archivos_variantes[0]

    if not guardar_imagen():
        return

    # Si hay más imágenes, son variantes
    if len(archivos_variantes) > 1:

        guardar_variantes(
            producto,
            archivos_variantes[1:]
        )

    # ------------------------------------------
    # ACTUALIZAR productos.json
    # ------------------------------------------

    generar_json()

    # ------------------------------------------
    # LIMPIAR ESTADO
    # ------------------------------------------

    archivo_seleccionado = None
    archivos_variantes = []

    btnConfirmar.config(
        state="disabled"
    )

    limpiar_previsualizacion()

    # Mostrar siguiente producto
    mostrar_producto()

    root.after(
        2000,
        lambda: actualizar_estado(
            "Seleccione una imagen para el siguiente producto."
        )
    )
# ==========================================
# OMITIR IMAGEN
# Asigna la imagen genérica y avanza al
# siguiente producto.
# ==========================================

def omitir_imagen():

    producto = obtener_producto_pendiente()

    if producto is None:
        return

    ws[producto["fila"]][5].value = "sin-imagen.png"
    wb.save(EXCEL)

    btnConfirmar.config(state="disabled")

    lblImagen.config(
        image="",
        text="Vista previa\n\n(Sin imagen)"
    )

    lblImagen.image = None

    mostrar_producto()

    actualizar_estado(
        "Producto omitido. Se asignó la imagen genérica."
    )

    root.after(
        2000,
        lambda: actualizar_estado(
            "Seleccione una imagen para el siguiente producto."
        )
    )
# ==========================================
# ACTUALIZAR ESTADO
# Muestra un mensaje en la barra de estado.
# ==========================================

def actualizar_estado(texto):

    subtitulo.config(text=texto)

    root.update_idletasks()

# ==========================================
# GEOMETRÍA DE LA VENTANA
# ==========================================

def cargar_geometria():

    if not RUTA_VENTANA_ASISTENTE.exists():
        return

    try:

        with open(
            RUTA_VENTANA_ASISTENTE,
            "r",
            encoding="utf-8"
        ) as archivo:

            datos = json.load(archivo)

        root.geometry(
            datos["geometry"]
        )

        if datos.get("state") == "zoomed":

            root.after(
                50,
                lambda: root.state("zoomed")
            )

    except Exception:
        pass


def guardar_geometria():

    try:

        RUTA_CONFIG.mkdir(
            exist_ok=True
        )

        datos = {
            "geometry": root.geometry(),
            "state": root.state()
        }

        with open(
            RUTA_VENTANA_ASISTENTE,
            "w",
            encoding="utf-8"
        ) as archivo:

            json.dump(
                datos,
                archivo,
                indent=4,
                ensure_ascii=False
            )

    except Exception:
        pass


def cerrar_aplicacion():

    guardar_geometria()

    root.destroy()
# ==========================================
# CREACIÓN DE LA VENTANA PRINCIPAL
# ==========================================

# Creamos la ventana principal de la aplicación.
root = tk.Tk()

root.protocol(
    "WM_DELETE_WINDOW",
    cerrar_aplicacion
)
try:
    root.iconbitmap(str(RUTA_ICONO))
except Exception:
    pass

root.after(
    100,
    lambda: root.focus_force()
)

# Título que aparecerá en la barra superior.
root.title("Asistente de imágenes - Catálogo Mi Mayo")

# Tamaño inicial de la ventana.

root.geometry("1150x850")

# Impide que la ventana sea demasiado pequeña.

root.minsize(900, 700)
# Color de fondo.
root.configure(bg="#f3f4f6")

# ==========================================
# TÍTULO PRINCIPAL
# ==========================================

titulo = tk.Label(
    root,
    text="Asistente de imágenes",
    font=("Segoe UI", 22, "bold"),
    bg="#f3f4f6",
)

titulo.pack(pady=(8, 4))

# ==========================================
# SUBTÍTULO
# ==========================================

subtitulo = tk.Label(
    root,
    text="",
    font=("Segoe UI", 11),
    fg="#2563EB",
    bg="#f3f4f6"
)

subtitulo.pack(pady=(0, 2))

# ==========================================
# LEYENDA DE USO
# ==========================================

instrucciones = tk.Label(
    root,
    text=(
        "¿Qué desea hacer?\n"
        "• Producto sin imagen → "
        "Seleccionar imagen y elegir una o varias imágenes.\n"
        "• Producto con imagen → "
        "Agregar variantes para incorporar imágenes adicionales.\n"
        "• Confirmar → guarda las imágenes seleccionadas.\n"
        "• Omitir → asigna la imagen genérica y continúa."
    ),
    font=("Segoe UI", 10),
    fg="#555",
    bg="#f3f4f6",
    justify="left",
    anchor="w",
    wraplength=1000
)

instrucciones.pack(
    padx=30,
    pady=(2, 4),
    anchor="w"
)
# ==========================================
# CONTENEDOR PRINCIPAL
# Agrupa todos los controles de la aplicación.
# ==========================================

frame = tk.Frame(
    root,
    bg="white",
    bd=1,
    relief="solid"
)

frame.pack(
    fill="both",
    expand=True,
    padx=25,
    pady=25
)

# ==========================================
# ESTRUCTURA PRINCIPAL DE LA APLICACIÓN
# ==========================================

# -------------------------------------------------
# Frame superior
# Contendrá:
#   • Imagen del producto (izquierda)
#   • Información del producto (derecha)
# -------------------------------------------------

frameSuperior = tk.Frame(frame, bg="white")

frameSuperior.pack(
    fill="x",
    padx=25,
    pady=20
)


# -------------------------------------------------
# Frame izquierdo
# Vista previa de la imagen
# -------------------------------------------------

frameImagen = tk.Frame(
    frameSuperior,
    bg="#f5f5f5",
    width=440,
    height=440,
    relief="solid",
    bd=1
)

frameImagen.pack(
    side="left",
    padx=(0,30)
)

frameImagen.pack_propagate(False)

# Texto temporal
lblImagen = tk.Label(
    frameImagen,
    text="Vista previa\n\n(Sin imagen)",
    bg="#f5f5f5",
    fg="#777",
    font=("Segoe UI",11)
)

lblImagen.pack(expand=True)

frameVariantes = tk.Frame(
    frameImagen,
    bg="#f5f5f5"
)

frameVariantes.pack(
    fill="x",
    padx=10,
    pady=10
)
# -------------------------------------------------
# Frame derecho
# Información del producto
# -------------------------------------------------

frameInfo = tk.Frame(
    frameSuperior,
    bg="white"
)

frameInfo.pack(
    side="left",
    anchor="n"
)

# Producto actual

lblProducto = tk.Label(
    frameInfo,
    text="Producto 0 de 0",
    font=("Segoe UI",18,"bold"),
    bg="white"
)

lblProducto.pack(anchor="w", pady=(0,20))

# Código

tk.Label(
    frameInfo,
    text="Código",
    font=("Segoe UI",10,"bold"),
    bg="white"
).pack(anchor="w")

lblCodigo = tk.Label(
    frameInfo,
    text="-",
    font=("Segoe UI",13),
    bg="white"
)

lblCodigo.pack(anchor="w", pady=(0,15))

# Nombre

tk.Label(
    frameInfo,
    text="Nombre",
    font=("Segoe UI",10,"bold"),
    bg="white"
).pack(anchor="w")

lblNombre = tk.Label(
    frameInfo,
    text="-",
    font=("Segoe UI",13),
    bg="white",
    wraplength=320,
    justify="left"
)

lblNombre.pack(anchor="w", pady=(0,15))

# Marca

tk.Label(
    frameInfo,
    text="Marca",
    font=("Segoe UI",10,"bold"),
    bg="white"
).pack(anchor="w")

lblMarca = tk.Label(
    frameInfo,
    text="-",
    font=("Segoe UI",13),
    bg="white"
)

lblMarca.pack(anchor="w", pady=(0,15))

# Categoría

tk.Label(
    frameInfo,
    text="Categoría",
    font=("Segoe UI",10,"bold"),
    bg="white"
).pack(anchor="w")

lblCategoria = tk.Label(
    frameInfo,
    text="-",
    font=("Segoe UI",13),
    bg="white"
)

lblCategoria.pack(anchor="w")

# ==========================================
# CONTENEDOR DE BOTONES
# 3 filas x 2 columnas
#
# Columna izquierda:
#   Seleccionar imagen
#   Agregar variantes
#   Administrar imágenes
#
# Columna derecha:
#   Confirmar
#   Guardar cambios
#   Omitir
#
# IMPORTANTE:
# No modifica frameImagen, frameVariantes,
# mostrar_vista_previa() ni el sistema de scroll.
# ==========================================

frameBotones = tk.Frame(
    frameInfo,
    bg="white"
)

frameBotones.pack(
    fill="x",
    pady=(15, 0)
)

for columna in range(2):
    frameBotones.grid_columnconfigure(
        columna,
        weight=1,
        uniform="boton"
    )

for fila in range(3):
    frameBotones.grid_rowconfigure(
        fila,
        weight=1
    )


# ==========================================
# BOTÓN SELECCIONAR IMAGEN
# ==========================================

btnSeleccionar = tk.Button(
    frameBotones,
    text="📂 Seleccionar imagen",
    font=("Segoe UI", 11, "bold"),
    bg="#2563EB",
    fg="white",
    padx=20,
    pady=10,
    cursor="hand2",
    command=seleccionar_imagen
)

btnSeleccionar.grid(
    row=0,
    column=0,
    sticky="ew",
    padx=(0, 5),
    pady=(0, 5)
)


# ==========================================
# BOTÓN CONFIRMAR
# Guarda definitivamente la imagen.
# ==========================================

btnConfirmar = tk.Button(
    frameBotones,
    text="✔ Confirmar",
    font=("Segoe UI", 11, "bold"),
    bg="#16A34A",
    fg="white",
    padx=20,
    pady=10,
    cursor="hand2",
    state="disabled",
    command=confirmar_imagen
)

btnConfirmar.grid(
    row=0,
    column=1,
    sticky="ew",
    padx=(5, 0),
    pady=(0, 5)
)


# ==========================================
# BOTÓN AGREGAR VARIANTES
# ==========================================

btnVariantes = tk.Button(
    frameBotones,
    text="🖼 Agregar variantes",
    font=("Segoe UI", 11, "bold"),
    bg="#7C3AED",
    fg="white",
    padx=20,
    pady=10,
    cursor="hand2",
    command=seleccionar_producto_variantes
)

btnVariantes.grid(
    row=1,
    column=0,
    sticky="ew",
    padx=(0, 5),
    pady=5
)


# ==========================================
# BOTÓN GUARDAR CAMBIOS
# Solo funciona para Administración.
# ==========================================

btnGuardarCambios = tk.Button(
    frameBotones,
    text="💾 Guardar cambios",
    font=("Segoe UI", 11, "bold"),
    bg="#16A34A",
    fg="white",
    padx=20,
    pady=10,
    cursor="hand2",
    state="disabled",
    command=guardar_cambios_imagenes
)

btnGuardarCambios.grid(
    row=1,
    column=1,
    sticky="ew",
    padx=(5, 0),
    pady=5
)


# ==========================================
# BOTÓN ADMINISTRAR IMÁGENES
# ==========================================

btnAdministrar = tk.Button(
    frameBotones,
    text="🛠 Administrar imágenes",
    font=("Segoe UI", 11, "bold"),
    bg="#0F766E",
    fg="white",
    padx=20,
    pady=10,
    cursor="hand2",
    command=administrar_imagenes
)

btnAdministrar.grid(
    row=2,
    column=0,
    sticky="ew",
    padx=(0, 5),
    pady=(5, 0)
)


# ==========================================
# BOTÓN OMITIR
# Asigna la imagen genérica al producto.
# ==========================================

btnOmitir = tk.Button(
    frameBotones,
    text="⏭ Omitir",
    font=("Segoe UI", 11, "bold"),
    bg="#F59E0B",
    fg="white",
    padx=20,
    pady=10,
    cursor="hand2",
    command=omitir_imagen
)

btnOmitir.grid(
    row=2,
    column=1,
    sticky="ew",
    padx=(5, 0),
    pady=(5, 0)
)

# ==========================================
# INICIAR LA APLICACIÓN
# ==========================================

def main():
    # Mostrar el primer producto pendiente
    mostrar_producto()
    actualizar_estado("Seleccione una imagen para comenzar.")
    cargar_geometria()
    root.mainloop()

if __name__ == "__main__":
    main()

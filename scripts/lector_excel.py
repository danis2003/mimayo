from openpyxl import load_workbook

from scripts.normalizadores import (
    normalizar_codigo,
    normalizar_texto,
    normalizar_precio,
    normalizar_imagen,
    normalizar_activo,
)

from scripts.config import HOJA_PRODUCTOS, RUTA_EXCEL

def leer_productos():
    libro = load_workbook(RUTA_EXCEL)
    hoja = libro[HOJA_PRODUCTOS]

    productos = []

    for fila_excel, fila in enumerate(
    hoja.iter_rows(min_row=2, values_only=True),
    start=2,
            ):

        codigo, nombre, marca, categoria, precio, imagen, activo, _ = fila

        codigo = normalizar_codigo(codigo)
        imagen = normalizar_imagen(imagen)
        nombre = normalizar_texto(nombre)
        marca = normalizar_texto(marca)
        categoria = normalizar_texto(categoria)

        producto = {
            "fila_excel": fila_excel,
            "codigo": codigo,
            "nombre": nombre,
            "marca": marca,
            "categoria": categoria,
            "precio": normalizar_precio(precio),
            "imagen": imagen,
            "activo": normalizar_activo(activo),
        }

        productos.append(producto)

    return productos
